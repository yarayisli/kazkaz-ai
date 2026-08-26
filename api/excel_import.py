"""Excel/CSV dosyalarını güvenli, açıklanabilir bir çalışma alanına dönüştürür."""

from __future__ import annotations

import csv
import io
import math
import re
import unicodedata
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Type

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ValidationError

from api.data_quality import kalite_raporu
from api.models import (
    AlacakFaturasi,
    BorcServisSatiri,
    ButceGerceklesmeSatiri,
    HaftalikNakitSatiri,
    MizanSatiri,
)
from api.transaction_analytics import islem_analizleri


MAKSIMUM_DOSYA = 5 * 1024 * 1024
MAKSIMUM_ACILMIS_XLSX = 80 * 1024 * 1024
MAKSIMUM_SATIR = 50_000


def _anahtar(deger: Any) -> str:
    metin = str(deger or "").strip().lower().replace("ı", "i")
    metin = unicodedata.normalize("NFKD", metin)
    metin = "".join(harf for harf in metin if not unicodedata.combining(harf))
    return re.sub(r"[^a-z0-9]+", "_", metin).strip("_")


ALANLAR = {
    "tarih": "tarih", "date": "tarih", "islem_tarihi": "tarih",
    "kategori": "kategori", "category": "kategori", "aciklama": "kategori",
    "gelir": "gelir", "income": "gelir", "tahsilat": "gelir",
    "gider": "gider", "expense": "gider", "odeme": "gider",
    "musteri": "musteri", "musteri_adi": "musteri", "customer": "musteri",
    "urun": "urun", "urun_adi": "urun", "product": "urun",
    "gider_tipi": "gider_tipi", "masraf_tipi": "gider_tipi",
    "vade_tarihi": "vade_tarihi", "vade": "vade_tarihi",
}

SAYFA_MODELLERI: Dict[str, Tuple[Type[BaseModel], Dict[str, str], str]] = {
    "mizan": (MizanSatiri, {
        "donem": "donem", "hesap_kodu": "hesap_kodu", "hesap_adi": "hesap_adi",
        "borc": "borc", "alacak": "alacak", "esleme": "esleme",
    }, "mizan"),
    "nakit_13_hafta": (HaftalikNakitSatiri, {
        "hafta": "hafta", "tahsilat": "tahsilat", "nakit_satis": "nakit_satis",
        "diger_giris": "diger_giris", "tedarikci": "tedarikci", "personel": "personel",
        "vergi": "vergi", "borc_servisi": "borc_servisi", "diger_cikis": "diger_cikis",
    }, "haftalik_nakit"),
    "alacaklar": (AlacakFaturasi, {
        "fatura_id": "fatura_id", "musteri_id": "musteri_id", "musteri_adi": "musteri_adi",
        "fatura_tarihi": "fatura_tarihi", "vade_tarihi": "vade_tarihi",
        "tutar": "tutar", "odenen": "odenen",
    }, "alacak_faturalari"),
    "borc_servisi": (BorcServisSatiri, {
        "borc_id": "borc_id", "alacakli": "alacakli", "odeme_tarihi": "odeme_tarihi",
        "anapara": "anapara", "faiz": "faiz", "para_birimi": "para_birimi",
    }, "borc_servisi"),
    "butce": (ButceGerceklesmeSatiri, {
        "ay": "ay", "kategori": "kategori", "departman": "departman", "proje": "proje",
        "butce": "butce", "gerceklesen": "gerceklesen", "onceki_tahmin": "onceki_tahmin",
    }, "butce"),
}


class DosyaIcerikHatasi(ValueError):
    pass


def veri_sablonu_olustur() -> bytes:
    """Desteklenen V1 veri sözleşmesini açıklayan boş ve güvenli Excel şablonu."""
    kitap = Workbook()
    rehber = kitap.active
    rehber.title = "Rehber"
    rehber.append(["KazKaz AI V1 Veri Şablonu"])
    rehber.append(["Tutarları pozitif girin; iade/düzeltmeleri negatif gelir veya gider olarak işaretleyin."])
    rehber.append(["Zorunlu olmayan sayfaları silebilirsiniz. Formül yerine değer yüklenmesi önerilir."])

    finans = kitap.create_sheet("Finansal_Gorunum")
    finans.append([
        "Şirket_Adı", "Sektör", "Dönem", "Ciro", "Satış_Maliyeti", "Faaliyet_Giderleri",
        "Net_Kâr", "Nakit", "Kısa_Vadeli_Borç", "Uzun_Vadeli_Borç", "Alacaklar",
        "Borçlar", "Stoklar", "Özkaynak", "Faiz_Gideri", "Vergi_Gideri", "Amortisman",
        "CapEx", "Dönen_Varlıklar", "Toplam_Varlıklar", "Toplam_Yükümlülükler",
        "Dağıtılmamış_Kârlar", "Operasyonel_Nakit_Akışı", "Dönem_Başı_Nakit",
        "Yatırım_Nakit_Akışı", "Finansman_Nakit_Akışı", "Dönem_Gün_Sayısı", "Etkin_Vergi_Oranı",
        "Rapor_Tarihi", "Minimum_Nakit_Eşiği",
    ])
    finans.append([
        "Örnek Şirket A.Ş.", "İmalat", "2026", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
        None, None, None, None, None, None, None, None, None, None, None, None, 365, None,
        None, None,
    ])

    islemler = kitap.create_sheet("İşlemler")
    islemler.append(["Tarih", "Kategori", "Gelir", "Gider", "Müşteri", "Ürün", "Gider_Tipi", "Vade_Tarihi"])
    islemler.append(["2026-01-01", "Satış", 0, 0, "", "", "", ""])

    sayfalar = {
        "Mizan": ["Dönem", "Hesap_Kodu", "Hesap_Adı", "Borç", "Alacak", "Eşleme"],
        "Nakit_13_Hafta": ["Hafta", "Tahsilat", "Nakit_Satış", "Diğer_Giriş", "Tedarikçi", "Personel", "Vergi", "Borç_Servisi", "Diğer_Çıkış"],
        "Alacaklar": ["Fatura_ID", "Müşteri_ID", "Müşteri_Adı", "Fatura_Tarihi", "Vade_Tarihi", "Tutar", "Ödenen"],
        "Borc_Servisi": ["Borç_ID", "Alacaklı", "Ödeme_Tarihi", "Anapara", "Faiz", "Para_Birimi"],
        "Butce": ["Ay", "Kategori", "Departman", "Proje", "Bütçe", "Gerçekleşen", "Önceki_Tahmin"],
    }
    for ad, basliklar in sayfalar.items():
        kitap.create_sheet(ad).append(basliklar)
    for sayfa in kitap.worksheets:
        sayfa.freeze_panes = "A2"
        sayfa.auto_filter.ref = sayfa.dimensions
    tampon = io.BytesIO()
    kitap.save(tampon)
    return tampon.getvalue()


def _sayi(deger: Any, varsayilan: Optional[float] = None) -> Optional[float]:
    if deger is None or (isinstance(deger, str) and not deger.strip()):
        return varsayilan
    if isinstance(deger, bool):
        raise ValueError("mantıksal değer sayı olarak kullanılamaz")
    if isinstance(deger, (int, float)):
        sonuc = float(deger)
    else:
        metin = str(deger).strip().replace("₺", "").replace("TRY", "").replace(" ", "")
        negatif = metin.startswith("(") and metin.endswith(")")
        metin = metin.strip("()")
        if "," in metin and "." in metin:
            if metin.rfind(",") > metin.rfind("."):
                metin = metin.replace(".", "").replace(",", ".")
            else:
                metin = metin.replace(",", "")
        elif "," in metin:
            parcalar = metin.split(",")
            metin = "".join(parcalar) if len(parcalar[-1]) == 3 else ".".join(parcalar)
        elif "." in metin and len(metin.rsplit(".", 1)[-1]) == 3:
            metin = metin.replace(".", "")
        sonuc = float(metin)
        if negatif:
            sonuc *= -1
    if not math.isfinite(sonuc):
        raise ValueError("sonlu olmayan sayı")
    return sonuc


def _tarih(deger: Any) -> Optional[date]:
    if deger is None or (isinstance(deger, str) and not deger.strip()):
        return None
    if isinstance(deger, datetime):
        return deger.date()
    if isinstance(deger, date):
        return deger
    metin = str(deger).strip()
    for bicim in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(metin, bicim).date()
        except ValueError:
            continue
    raise ValueError("desteklenmeyen tarih biçimi")


def _json_degeri(deger: Any) -> Any:
    if isinstance(deger, (date, datetime)):
        return deger.isoformat()
    return deger


def _hata(hatalar: List[Dict[str, Any]], sayfa: str, satir: int, alan: str, mesaj: str,
          seviye: str = "hata", kod: str = "gecersiz_deger") -> None:
    if len(hatalar) < 500:
        hatalar.append({"sayfa": sayfa, "satir": satir, "alan": alan, "kod": kod,
                        "mesaj": mesaj, "seviye": seviye})


def _xlsx_guvenlik_kontrolu(icerik: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(icerik)) as arsiv:
            uyeler = arsiv.infolist()
            acilmis = sum(uye.file_size for uye in uyeler)
            if len(uyeler) > 3_000 or acilmis > MAKSIMUM_ACILMIS_XLSX:
                raise DosyaIcerikHatasi("Excel arşivi güvenli açılma sınırını aşıyor.")
            for uye in uyeler:
                if uye.compress_size and uye.file_size > 10 * 1024 * 1024:
                    if uye.file_size / uye.compress_size > 200:
                        raise DosyaIcerikHatasi("Şüpheli derecede sıkıştırılmış Excel içeriği reddedildi.")
    except zipfile.BadZipFile as exc:
        raise DosyaIcerikHatasi("Dosya geçerli bir .xlsx çalışma kitabı değil.") from exc


def _baslik_satiri(satirlar: Iterable[Tuple[Any, ...]], bilinen: set[str]) -> Tuple[int, Dict[int, str]]:
    en_iyi: Tuple[int, int, Dict[int, str]] = (-1, 0, {})
    for sira, satir in enumerate(satirlar, start=1):
        esleme = {indeks: _anahtar(deger) for indeks, deger in enumerate(satir) if deger is not None}
        puan = sum(1 for alan in esleme.values() if alan in bilinen)
        if puan > en_iyi[1]:
            en_iyi = (sira, puan, esleme)
        if sira >= 15:
            break
    if en_iyi[1] < 2:
        raise DosyaIcerikHatasi("Tanınan sütun başlıkları bulunamadı.")
    return en_iyi[0], en_iyi[2]


def _model_sayfasi(ws: Any, sayfa_adi: str, model: Type[BaseModel], alanlar: Dict[str, str],
                   hedef: List[Dict[str, Any]], hatalar: List[Dict[str, Any]],
                   onizleme: List[Dict[str, Any]]) -> Tuple[int, int]:
    baslik_no, basliklar = _baslik_satiri(ws.iter_rows(values_only=True), set(alanlar))
    gecerli = reddedilen = 0
    for satir_no, hucreler in enumerate(ws.iter_rows(min_row=baslik_no + 1, values_only=True), start=baslik_no + 1):
        if not any(deger not in (None, "") for deger in hucreler):
            continue
        ham: Dict[str, Any] = {}
        for indeks, deger in enumerate(hucreler):
            kaynak = basliklar.get(indeks)
            if kaynak in alanlar:
                ham[alanlar[kaynak]] = deger
        try:
            kayit = model.model_validate(ham).model_dump(mode="json")
            hedef.append(kayit)
            gecerli += 1
            if len(onizleme) < 30:
                onizleme.append({"sayfa": sayfa_adi, "satir": satir_no, **kayit})
        except ValidationError as exc:
            reddedilen += 1
            for ayrinti in exc.errors()[:4]:
                konum = ayrinti.get("loc") or ("satir",)
                _hata(hatalar, sayfa_adi, satir_no, str(konum[0]), ayrinti["msg"])
    return gecerli, reddedilen


def _mukerrerleri_ayikla(
    kayitlar: List[Dict[str, Any]], alan: str, sayfa: str, hatalar: List[Dict[str, Any]],
) -> int:
    """Kimlik alanı tekrarlanan kayıtları ilk satırı koruyarak reddeder."""
    gorulen = set()
    benzersiz = []
    reddedilen = 0
    for indeks, kayit in enumerate(kayitlar, start=2):
        kimlik = str(kayit.get(alan) or "").strip()
        if kimlik and kimlik in gorulen:
            reddedilen += 1
            _hata(hatalar, sayfa, indeks, alan, f"Tekrarlanan kimlik reddedildi: {kimlik}", "hata", "mukerrer_kimlik")
            continue
        gorulen.add(kimlik)
        benzersiz.append(kayit)
    kayitlar[:] = benzersiz
    return reddedilen


def _islem_satirlari(satirlar: List[List[Any]], sayfa_adi: str, hatalar: List[Dict[str, Any]],
                     onizleme: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], int]:
    baslik_no, basliklar = _baslik_satiri(iter(tuple(s) for s in satirlar[:15]), set(ALANLAR))
    standart = {indeks: ALANLAR.get(alan) for indeks, alan in basliklar.items()}
    zorunlu = set(standart.values())
    if not {"tarih", "kategori", "gelir", "gider"}.issubset(zorunlu):
        raise DosyaIcerikHatasi("İşlem dosyasında Tarih, Kategori, Gelir ve Gider sütunları zorunludur.")
    sonuc: List[Dict[str, Any]] = []
    reddedilen = 0
    for satir_no, hucreler in enumerate(satirlar[baslik_no:], start=baslik_no + 1):
        if len(sonuc) >= MAKSIMUM_SATIR:
            _hata(hatalar, sayfa_adi, satir_no, "satir", "50.000 satır sınırından sonrası alınmadı.", "uyari", "satir_siniri")
            break
        if not any(deger not in (None, "") for deger in hucreler):
            continue
        ham = {alan: hucreler[indeks] if indeks < len(hucreler) else None
               for indeks, alan in standart.items() if alan}
        try:
            tarih = _tarih(ham.get("tarih"))
            kategori = str(ham.get("kategori") or "").strip()
            if not tarih or not kategori:
                raise ValueError("Tarih ve kategori boş bırakılamaz")
            kayit = {
                "tarih": tarih.isoformat(), "kategori": kategori,
                "gelir": _sayi(ham.get("gelir"), 0) or 0,
                "gider": _sayi(ham.get("gider"), 0) or 0,
            }
            for alan in ("musteri", "urun", "gider_tipi"):
                if ham.get(alan) not in (None, ""):
                    kayit[alan] = str(ham[alan]).strip()
            if ham.get("vade_tarihi") not in (None, ""):
                kayit["vade_tarihi"] = _tarih(ham["vade_tarihi"]).isoformat()
            if kayit["gelir"] < 0 or kayit["gider"] < 0:
                _hata(hatalar, sayfa_adi, satir_no, "tutar", "Negatif tutar iade/düzeltme olarak korundu.", "uyari", "negatif_tutar")
            sonuc.append(kayit)
            if len(onizleme) < 30:
                onizleme.append({"sayfa": sayfa_adi, "satir": satir_no, **kayit})
        except (ValueError, TypeError) as exc:
            reddedilen += 1
            _hata(hatalar, sayfa_adi, satir_no, "satir", str(exc))
    return sonuc, reddedilen


def _finansal_ozet(islemler: List[Dict[str, Any]]) -> Dict[str, Any]:
    gelir = sum(s["gelir"] for s in islemler)
    gider = sum(s["gider"] for s in islemler)
    cogs = faaliyet = faiz = vergi = amortisman = capex = 0.0
    for satir in islemler:
        anahtar = _anahtar(f"{satir.get('kategori', '')} {satir.get('gider_tipi', '')}")
        tutar = satir["gider"]
        if any(k in anahtar for k in ("hammadde", "malzeme", "satis_maliyeti", "uretim_maliyeti", "cogs", "smm")):
            cogs += tutar
        elif "faiz" in anahtar:
            faiz += tutar
        elif "vergi" in anahtar:
            vergi += tutar
        elif any(k in anahtar for k in ("amortisman", "itfa")):
            amortisman += tutar
        elif any(k in anahtar for k in ("capex", "yatirim_harcamasi", "demirbas")):
            capex += tutar
        else:
            faaliyet += tutar
    tarihler = sorted(s["tarih"] for s in islemler)
    donem = f"{tarihler[0]} / {tarihler[-1]}" if tarihler else "Güncel"
    return {
        "sirket_adi": "İçe Aktarılan Şirket", "sektor": "Belirtilmedi", "donem": donem,
        "ciro": max(0, gelir), "satis_maliyeti": max(0, cogs),
        "faaliyet_giderleri": max(0, faaliyet), "net_kar": gelir - gider,
        "nakit": 0, "kisa_vadeli_borc": 0, "uzun_vadeli_borc": 0,
        "alacaklar": 0, "borclar": 0, "stoklar": 0, "ozkaynak": 0,
        "faiz_gideri": max(0, faiz), "vergi_gideri": max(0, vergi),
        "amortisman": max(0, amortisman), "capex": max(0, capex),
    }


def _finansal_gorunum(ws: Any, hatalar: List[Dict[str, Any]], onizleme: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    alanlar = {
        "sirket_adi": "sirket_adi", "sektor": "sektor", "donem": "donem", "ciro": "ciro",
        "satis_maliyeti": "satis_maliyeti", "faaliyet_giderleri": "faaliyet_giderleri",
        "net_kar": "net_kar", "nakit": "nakit", "kisa_vadeli_borc": "kisa_vadeli_borc",
        "uzun_vadeli_borc": "uzun_vadeli_borc", "alacaklar": "alacaklar", "borclar": "borclar",
        "stoklar": "stoklar", "ozkaynak": "ozkaynak", "faiz_gideri": "faiz_gideri",
        "vergi_gideri": "vergi_gideri", "amortisman": "amortisman", "capex": "capex",
        "donen_varliklar": "donen_varliklar", "toplam_varliklar": "toplam_varliklar",
        "toplam_yukumlulukler": "toplam_yukumlulukler",
        "dagitilmamis_karlar": "dagitilmamis_karlar",
        "operasyonel_nakit_akisi": "operasyonel_nakit_akisi",
        "donem_basi_nakit": "donem_basi_nakit",
        "yatirim_nakit_akisi": "yatirim_nakit_akisi",
        "finansman_nakit_akisi": "finansman_nakit_akisi",
        "donem_gun_sayisi": "donem_gun_sayisi", "etkin_vergi_orani": "etkin_vergi_orani",
        "rapor_tarihi": "rapor_tarihi", "minimum_nakit_esigi": "minimum_nakit_esigi",
    }
    baslik_no, basliklar = _baslik_satiri(ws.iter_rows(values_only=True), set(alanlar))
    for satir_no, hucreler in enumerate(ws.iter_rows(min_row=baslik_no + 1, values_only=True), start=baslik_no + 1):
        ham = {alanlar[basliklar[i]]: deger for i, deger in enumerate(hucreler)
               if i in basliklar and basliklar[i] in alanlar}
        if not ham or ham.get("ciro") in (None, ""):
            continue
        try:
            metinler = {"sirket_adi", "sektor", "donem"}
            opsiyonel_sayilar = {
                "faiz_gideri", "vergi_gideri", "amortisman", "capex", "donen_varliklar",
                "toplam_varliklar", "toplam_yukumlulukler", "dagitilmamis_karlar",
                "operasyonel_nakit_akisi", "donem_basi_nakit", "yatirim_nakit_akisi",
                "finansman_nakit_akisi", "donem_gun_sayisi", "etkin_vergi_orani",
                "minimum_nakit_esigi",
            }
            kayit = {
                alan: (
                    _tarih(deger).isoformat()
                    if alan == "rapor_tarihi" and _tarih(deger) is not None
                    else str(deger).strip()
                    if alan in metinler
                    else _sayi(deger, None if alan in opsiyonel_sayilar else 0)
                )
                for alan, deger in ham.items()
            }
            kayit.setdefault("sirket_adi", "İçe Aktarılan Şirket")
            kayit.setdefault("sektor", "Belirtilmedi")
            kayit.setdefault("donem", "Güncel")
            for alan in ("satis_maliyeti", "faaliyet_giderleri", "net_kar", "nakit", "kisa_vadeli_borc",
                         "uzun_vadeli_borc", "alacaklar", "borclar", "stoklar", "ozkaynak"):
                kayit.setdefault(alan, 0)
            onizleme.append({"sayfa": ws.title, "satir": satir_no, **kayit})
            return kayit
        except ValueError as exc:
            _hata(hatalar, ws.title, satir_no, "finansal_gorunum", str(exc))
    return None


def _csv_satirlari(icerik: bytes) -> List[List[Any]]:
    metin = None
    for kodlama in ("utf-8-sig", "utf-8", "cp1254", "latin-1"):
        try:
            metin = icerik.decode(kodlama)
            break
        except UnicodeDecodeError:
            continue
    if metin is None:
        raise DosyaIcerikHatasi("CSV karakter kodlaması okunamadı.")
    ornek = metin[:8192]
    try:
        lehce = csv.Sniffer().sniff(ornek, delimiters=",;\t|")
    except csv.Error:
        lehce = csv.excel
    return [list(satir) for satir in csv.reader(io.StringIO(metin), dialect=lehce)]


def dosya_dogrula(icerik: bytes, dosya_adi: str) -> Dict[str, Any]:
    """Kullanıcı dosyasını çalıştırmadan okur, doğrular ve tek veri sözleşmesine çevirir."""
    if not icerik:
        raise DosyaIcerikHatasi("Dosya boş.")
    if len(icerik) > MAKSIMUM_DOSYA:
        raise DosyaIcerikHatasi("Dosya boyutu 5 MB sınırını aşıyor.")
    temiz_ad = Path(dosya_adi).name[:180]
    uzanti = Path(temiz_ad).suffix.lower()
    if uzanti not in {".xlsx", ".csv"}:
        raise DosyaIcerikHatasi("Yalnızca .xlsx ve .csv dosyaları kabul edilir; eski .xls dosyasını .xlsx olarak kaydedin.")

    hatalar: List[Dict[str, Any]] = []
    onizleme: List[Dict[str, Any]] = []
    gelismis: Dict[str, Any] = {
        "rapor_tarihi": date.today().isoformat(), "baslangic_nakdi": 0,
        "mizan": [], "haftalik_nakit": [], "alacak_faturalari": [], "borc_servisi": [], "butce": [],
    }
    islemler: List[Dict[str, Any]] = []
    finansal: Optional[Dict[str, Any]] = None
    finansal_sayfa_var = False
    gecerli = reddedilen = 0
    sayfalar: List[str] = []

    if uzanti == ".csv":
        sayfalar = ["CSV"]
        islemler, reddedilen = _islem_satirlari(_csv_satirlari(icerik), "CSV", hatalar, onizleme)
        gecerli = len(islemler)
    else:
        _xlsx_guvenlik_kontrolu(icerik)
        try:
            kitap = load_workbook(io.BytesIO(icerik), read_only=True, data_only=True)
        except Exception as exc:
            raise DosyaIcerikHatasi("Excel çalışma kitabı açılamadı.") from exc
        sayfalar = kitap.sheetnames
        if "Finansal_Gorunum" in kitap.sheetnames:
            finansal = _finansal_gorunum(kitap["Finansal_Gorunum"], hatalar, onizleme)
            finansal_sayfa_var = finansal is not None
            gecerli += 1 if finansal else 0
        for sayfa_adi, (model, alanlar, hedef_adi) in SAYFA_MODELLERI.items():
            gercek_ad = next((ad for ad in kitap.sheetnames if _anahtar(ad) == sayfa_adi), None)
            if not gercek_ad:
                continue
            eklenen, atlanan = _model_sayfasi(
                kitap[gercek_ad], gercek_ad, model, alanlar, gelismis[hedef_adi], hatalar, onizleme,
            )
            gecerli += eklenen
            reddedilen += atlanan
        reddedilen += _mukerrerleri_ayikla(gelismis["alacak_faturalari"], "fatura_id", "Alacaklar", hatalar)
        # İşlem sayfasını adıyla önceliklendir. Kontrol/rehber sayfalarının sırası
        # değişse bile yanlışlıkla işlem verisi kabul edilmemelidir.
        aday = next(
            (ad for ad in kitap.sheetnames if _anahtar(ad) in {"islemler", "transactions"}),
            None,
        )
        bilinen = {
            "finansal_gorunum", "islemler", "transactions", *SAYFA_MODELLERI.keys(),
            "kapak", "rehber", "kontroller", "kontrol_paneli", "beklenen_sonuclar",
            "test_senaryolari", "hacim_5000",
        }
        if not aday:
            aday = next((ad for ad in kitap.sheetnames if _anahtar(ad) not in bilinen), None)
        if aday:
            ham_satirlar = [list(s) for s in kitap[aday].iter_rows(values_only=True)]
            try:
                islemler, atlanan = _islem_satirlari(ham_satirlar, aday, hatalar, onizleme)
                gecerli += len(islemler)
                reddedilen += atlanan
            except DosyaIcerikHatasi:
                pass
        kitap.close()

    if not finansal and islemler:
        finansal = _finansal_ozet(islemler)
        _hata(hatalar, "Genel", 0, "bilanço", "İşlem dosyasında bilanço yok; bilanço alanları sıfır bırakıldı.", "uyari", "eksik_bilanco")
    if not finansal:
        raise DosyaIcerikHatasi("Finansal görünüm veya Tarih/Kategori/Gelir/Gider işlemleri bulunamadı.")

    rapor_tarihi = finansal.pop("rapor_tarihi", None)
    minimum_nakit_esigi = finansal.pop("minimum_nakit_esigi", None)
    # 13 haftalık projeksiyon dönem sonu nakdiyle değil, kullanıcı tarafından
    # verilen dönem başı bakiyesiyle başlamalıdır. Alan yoksa geriye dönük
    # uyumluluk için mevcut nakit bakiyesine düşer.
    gelismis["baslangic_nakdi"] = (
        finansal.get("donem_basi_nakit")
        if finansal.get("donem_basi_nakit") is not None
        else finansal.get("nakit", 0)
    )
    gelismis["operasyonel_nakit_akisi"] = finansal.get("operasyonel_nakit_akisi")
    if minimum_nakit_esigi is not None:
        gelismis["minimum_nakit_esigi"] = minimum_nakit_esigi
    if rapor_tarihi:
        gelismis["rapor_tarihi"] = rapor_tarihi
    elif gelismis["alacak_faturalari"]:
        gelismis["rapor_tarihi"] = max(s["fatura_tarihi"] for s in gelismis["alacak_faturalari"])
    analizler = islem_analizleri(islemler)
    if analizler.get("musteriler"):
        finansal["musteri_cirolari"] = [
            {"musteri_id": satir["id"], "musteri_adi": satir["ad"], "ciro": satir["gelir"]}
            for satir in analizler["musteriler"]
            if satir.get("gelir", 0) > 0
        ]
    toplam_gelir = sum(s["gelir"] for s in islemler)
    toplam_gider = sum(s["gider"] for s in islemler)
    uyari_sayisi = sum(1 for h in hatalar if h["seviye"] == "uyari")

    # Sayfa eşleme özeti — hangi sayfalar tanındı, hangileri atlandı.
    tanınan_sayfalar: List[str] = []
    atlanan_sayfalar: List[str] = []
    if uzanti != ".csv":
        _bilinen_sayfalar = {
            "finansal_gorunum", "islemler", "transactions", *SAYFA_MODELLERI.keys(),
            "kapak", "rehber", "kontroller", "kontrol_paneli",
            "beklenen_sonuclar", "test_senaryolari", "hacim_5000",
        }
        for ad in sayfalar:
            anahtar = _anahtar(ad)
            if anahtar in _bilinen_sayfalar:
                tanınan_sayfalar.append(ad)
            else:
                atlanan_sayfalar.append(ad)

    # Semantik veri kalitesi raporu (cross-field + anomali)
    kalite = kalite_raporu(
        finansal_veri=finansal,
        zaman_serisi=islemler,
        musteri_cirolari=finansal.get("musteri_cirolari") if finansal else None,
    )

    return {
        "durum": "hazir" if not hatalar else "uyarili",
        "dosya": {
            "ad": temiz_ad, "tur": uzanti[1:], "boyut": len(icerik),
            "sayfalar": sayfalar,
            "tanınan_sayfalar": tanınan_sayfalar,
            "atlanan_sayfalar": atlanan_sayfalar,
        },
        "ozet": {
            "gecerli_satirlar": gecerli, "uyarili_satirlar": uyari_sayisi,
            "reddedilen_satirlar": reddedilen, "toplam_gelir": toplam_gelir,
            "toplam_gider": toplam_gider, "islem_satirlari": len(islemler),
        },
        "finansal_veri": finansal,
        "veri_kalitesi": {
            "kaynak": "finansal_gorunum" if finansal_sayfa_var else "islem_ozeti",
            "bilanco_mevcut": finansal_sayfa_var,
            "favok_hesaplanabilir": all(
                finansal.get(alan) is not None
                for alan in ("faiz_gideri", "vergi_gideri", "amortisman")
            ),
            "kurumsal_metrikler_hazir": all(
                finansal.get(alan) is not None
                for alan in ("donen_varliklar", "toplam_varliklar", "toplam_yukumlulukler", "dagitilmamis_karlar")
            ),
            "eksikler": [] if finansal_sayfa_var else [
                "Bilanço kalemleri", "Faiz/vergi/amortisman ayrımının açık teyidi",
            ],
            # Cross-field tutarlılık + anomali bulguları — kullanıcı Uygula
            # butonuna basmadan önce görebilir. Bulgular hesaplamayı bloke
            # etmez; sadece dikkat çeker.
            "tutarlilik_bulgulari": kalite["tutarlilik_bulgulari"],
            "anomali_bulgulari": kalite["anomali_bulgulari"],
            "semantik_durum": kalite["durum"],  # temiz | uyarili | hatali
            "semantik_hata_sayisi": kalite["toplam_hata"],
            "semantik_uyari_sayisi": kalite["toplam_uyari"],
        },
        "gelismis_veri": gelismis,
        "zaman_serisi": islemler,
        "analizler": analizler,
        "onizleme": onizleme[:30],
        "hatalar": hatalar,
        "metodoloji": {
            "negatif_tutarlar": "İade/düzeltme olarak korunur.",
            "favok": "Net kâr + faiz + vergi + amortisman; yalnızca ayrıştırılmış alanlarla.",
            "ai": "Dosya okuma ve hesaplama deterministiktir; AI veri ayrıştırmasında kullanılmaz.",
        },
    }
