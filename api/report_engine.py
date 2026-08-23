"""KazKaz V1 yönetici PDF ve Excel rapor üretimi."""

from __future__ import annotations

import io
import os
from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from api.models import FinansalGorunum
from api.services import finansal_denetim


LACIVERT = colors.HexColor("#0F2252")
TURUNCU = colors.HexColor("#FF4D00")
ACIK = colors.HexColor("#F3F6FA")


def _font() -> str:
    adaylar = [
        os.getenv("REPORT_FONT_PATH", ""),
        str(Path(__file__).resolve().parents[1] / "assets" / "DejaVuSans.ttf"),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ]
    for aday in adaylar:
        if aday and Path(aday).exists():
            if "KazKazUnicode" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("KazKazUnicode", aday))
            return "KazKazUnicode"
    return "Helvetica"


def _para(deger: Any, para_birimi: str = "TRY") -> str:
    if deger is None:
        return "Hesaplanamadı"
    return f"{float(deger):,.2f} {para_birimi}"


def _oran(deger: Any) -> str:
    if deger is None:
        return "Hesaplanamadı"
    return f"%{float(deger):,.2f}"


def _rapor(veri: FinansalGorunum) -> Dict[str, Any]:
    return finansal_denetim(veri)


def pdf_raporu_olustur(veri: FinansalGorunum) -> bytes:
    rapor = _rapor(veri)
    tampon = io.BytesIO()
    font = _font()
    stiller = getSampleStyleSheet()
    baslik = ParagraphStyle("Baslik", parent=stiller["Title"], fontName=font, fontSize=20, leading=24, textColor=LACIVERT, alignment=TA_LEFT)
    alt = ParagraphStyle("Alt", parent=stiller["BodyText"], fontName=font, fontSize=9, leading=13, textColor=colors.HexColor("#526078"))
    bolum = ParagraphStyle("Bolum", parent=stiller["Heading2"], fontName=font, fontSize=12, leading=16, textColor=LACIVERT, spaceBefore=10, spaceAfter=7)
    govde = ParagraphStyle("Govde", parent=stiller["BodyText"], fontName=font, fontSize=9, leading=13)

    belge = SimpleDocTemplate(
        tampon,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"{veri.sirket_adi} - KazKaz AI Yönetici Raporu",
        author="KazKaz AI",
    )
    akis = [
        Paragraph("KazKaz AI Yönetici Finans Raporu", baslik),
        Paragraph(f"{veri.sirket_adi} · {veri.donem} · Karar destek raporu", alt),
        Spacer(1, 7 * mm),
    ]

    metrikler = rapor["metrikler"]
    tablo_verisi = [
        ["Gösterge", "Sonuç"],
        ["Ciro", _para(veri.ciro, veri.para_birimi)],
        ["Net kâr", _para(veri.net_kar, veri.para_birimi)],
        ["Net kâr marjı", _oran(metrikler["net_kar_marji"])],
        ["FAVÖK", _para(metrikler["favok"], veri.para_birimi)],
        ["Cari oran", "Hesaplanamadı" if metrikler["cari_oran"] is None else f"{metrikler['cari_oran']:.2f}"],
        ["Serbest nakit akışı", _para(metrikler["serbest_nakit_akisi"], veri.para_birimi)],
        ["Altman Z'", "Hesaplanamadı" if metrikler["altman_z_prime"] is None else f"{metrikler['altman_z_prime']:.2f}"],
    ]
    tablo = Table(tablo_verisi, colWidths=[105 * mm, 55 * mm], repeatRows=1)
    tablo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), LACIVERT),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 1), (-1, -1), ACIK),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6DCE7")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    akis.extend([tablo, Paragraph("Riskler", bolum)])
    riskler = rapor.get("riskler") or ["Doğrulanmış kritik risk bulunmadı."]
    akis.extend(Paragraph(f"• {risk}", govde) for risk in riskler)
    akis.append(Paragraph("Önerilen aksiyonlar", bolum))
    aksiyonlar = rapor.get("aksiyonlar") or ["Veri kapsamını genişletin ve dönemsel karşılaştırmayı sürdürün."]
    akis.extend(Paragraph(f"• {aksiyon}", govde) for aksiyon in aksiyonlar)
    akis.extend([
        Paragraph("Veri kalitesi", bolum),
        Paragraph(
            f"Seviye: {rapor['veri_kalitesi']['seviye']} · Skor: {rapor['veri_kalitesi']['skor']}/100. "
            "Eksik verilerde sistem değer üretmez; sonuçlar karar desteğidir.",
            govde,
        ),
        Spacer(1, 5 * mm),
        Paragraph(rapor["uyari"], alt),
    ])
    belge.build(akis)
    return tampon.getvalue()


def excel_raporu_olustur(veri: FinansalGorunum) -> bytes:
    rapor = _rapor(veri)
    kitap = Workbook()
    ozet = kitap.active
    ozet.title = "Yönetici Özeti"
    ozet.append(["KazKaz AI Yönetici Finans Raporu"])
    ozet.append(["Şirket", veri.sirket_adi])
    ozet.append(["Dönem", veri.donem])
    ozet.append(["Raporlama Para Birimi", veri.para_birimi])
    ozet.append(["Ciro", veri.ciro])
    ozet.append(["Net Kâr", veri.net_kar])
    ozet.append(["Veri Kalitesi", rapor["veri_kalitesi"]["seviye"]])
    ozet.append(["Veri Kalitesi Skoru", rapor["veri_kalitesi"]["skor"]])

    metrik = kitap.create_sheet("Metrikler")
    metrik.append(["Metrik", "Değer", "Durum", "Formül", "Güven", "Eksik Alanlar"])
    for ad, kayit in rapor["metrik_kaydi"].items():
        metrik.append([
            ad,
            kayit["deger"],
            kayit["durum"],
            kayit["formula"],
            kayit["guven"],
            ", ".join(kayit["eksik_alanlar"]),
        ])

    bulgular = kitap.create_sheet("Risk ve Aksiyonlar")
    bulgular.append(["Tür", "Açıklama"])
    for risk in rapor.get("riskler", []):
        bulgular.append(["Risk", risk])
    for aksiyon in rapor.get("aksiyonlar", []):
        bulgular.append(["Aksiyon", aksiyon])

    for sayfa in kitap.worksheets:
        sayfa.freeze_panes = "A2"
        sayfa.auto_filter.ref = sayfa.dimensions
        sayfa.column_dimensions["A"].width = 28
        for sutun in "BCDEF":
            sayfa.column_dimensions[sutun].width = 24
        for hucre in sayfa[1]:
            hucre.fill = PatternFill("solid", fgColor="0F2252")
            hucre.font = Font(color="FFFFFF", bold=True)
            hucre.alignment = Alignment(vertical="center")

    tampon = io.BytesIO()
    kitap.save(tampon)
    return tampon.getvalue()
