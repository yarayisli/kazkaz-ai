import io
import os
import unittest
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook

from api.excel_import import DosyaIcerikHatasi, dosya_dogrula, veri_sablonu_olustur
from api.main import uygulama
from api.security_middleware import hiz_limitlerini_sifirla


class TestDosyaIceriAktarma(unittest.TestCase):
    def setUp(self):
        hiz_limitlerini_sifirla()

    def test_csv_turkce_sayilari_ve_negatif_iadeyi_korur(self):
        icerik = (
            "Tarih;Kategori;Gelir;Gider;Müşteri;Ürün\n"
            "01.01.2026;Satış;1.250,50;0;Müşteri A;Ürün 1\n"
            "02.01.2026;Satış İadesi;-250,50;0;Müşteri A;Ürün 1\n"
            "03.01.2026;Hammadde;0;400,00;;\n"
        ).encode("utf-8")
        sonuc = dosya_dogrula(icerik, "ornek.csv")

        self.assertEqual(sonuc["ozet"]["islem_satirlari"], 3)
        self.assertEqual(sonuc["ozet"]["toplam_gelir"], 1000)
        self.assertEqual(sonuc["finansal_veri"]["satis_maliyeti"], 400)
        self.assertTrue(any(h["kod"] == "negatif_tutar" for h in sonuc["hatalar"]))

    def test_calisma_kitabi_tum_finans_koleksiyonlarini_uretir(self):
        kitap = Workbook()
        finans = kitap.active
        finans.title = "Finansal_Gorunum"
        finans.append(["KazKaz test paketi"])
        finans.append([])
        finans.append(["Şirket_Adı", "Sektör", "Dönem", "Ciro", "Satış_Maliyeti", "Faaliyet_Giderleri", "Net_Kâr", "Nakit", "Kısa_Vadeli_Borç", "Uzun_Vadeli_Borç", "Alacaklar", "Borçlar", "Stoklar", "Özkaynak", "Faiz_Gideri", "Vergi_Gideri", "Amortisman", "CapEx", "Dönen_Varlıklar", "Toplam_Varlıklar", "Toplam_Yükümlülükler", "Dağıtılmamış_Kârlar", "Operasyonel_Nakit_Akışı", "Dönem_Başı_Nakit", "Yatırım_Nakit_Akışı", "Finansman_Nakit_Akışı", "Dönem_Gün_Sayısı", "Etkin_Vergi_Oranı", "Rapor_Tarihi", "Minimum_Nakit_Eşiği"])
        finans.append(["Test AŞ", "Yazılım", "2026-07", 1_000_000, 200_000, 300_000, 350_000, 100_000, 80_000, 120_000, 150_000, 60_000, 20_000, 500_000, 50_000, 30_000, 40_000, 70_000, 350_000, 900_000, 400_000, 180_000, 420_000, 80_000, -300_000, -100_000, 212, 20, "2026-07-31", 75_000])

        mizan = kitap.create_sheet("Mizan")
        mizan.append(["Dönem", "Hesap_Kodu", "Hesap_Adı", "Borç", "Alacak", "Eşleme"])
        mizan.append(["2026-07", "100", "Kasa", 100_000, 0, "nakit"])

        nakit = kitap.create_sheet("Nakit_13_Hafta")
        nakit.append(["Hafta", "Tahsilat", "Nakit_Satış", "Diğer_Giriş", "Tedarikçi", "Personel", "Vergi", "Borç_Servisi", "Diğer_Çıkış"])
        nakit.append(["2026-08-17", 20_000, 5_000, 0, 7_000, 6_000, 1_000, 2_000, 500])

        tampon = io.BytesIO()
        kitap.save(tampon)
        sonuc = dosya_dogrula(tampon.getvalue(), "tam_paket.xlsx")

        self.assertEqual(sonuc["finansal_veri"]["sirket_adi"], "Test AŞ")
        self.assertEqual(sonuc["finansal_veri"]["toplam_varliklar"], 900_000)
        self.assertEqual(sonuc["finansal_veri"]["operasyonel_nakit_akisi"], 420_000)
        self.assertEqual(sonuc["finansal_veri"]["donem_basi_nakit"], 80_000)
        self.assertEqual(sonuc["finansal_veri"]["yatirim_nakit_akisi"], -300_000)
        self.assertEqual(sonuc["finansal_veri"]["finansman_nakit_akisi"], -100_000)
        self.assertEqual(sonuc["gelismis_veri"]["rapor_tarihi"], "2026-07-31")
        self.assertEqual(sonuc["gelismis_veri"]["minimum_nakit_esigi"], 75_000)
        self.assertEqual(sonuc["gelismis_veri"]["operasyonel_nakit_akisi"], 420_000)
        self.assertEqual(sonuc["gelismis_veri"]["baslangic_nakdi"], 80_000)
        self.assertTrue(sonuc["veri_kalitesi"]["kurumsal_metrikler_hazir"])
        self.assertEqual(len(sonuc["gelismis_veri"]["mizan"]), 1)
        self.assertEqual(len(sonuc["gelismis_veri"]["haftalik_nakit"]), 1)

    def test_xls_ve_calistirilabilir_dosya_reddedilir(self):
        for ad in ("eski.xls", "zararli.xlsm", "program.exe"):
            with self.subTest(ad=ad), self.assertRaises(DosyaIcerikHatasi):
                dosya_dogrula(b"ornek", ad)

    def test_api_ham_dosya_govdesini_kabul_eder(self):
        istemci = TestClient(uygulama)
        with patch.dict(os.environ, {"KAZKAZ_AUTH_DISABLED": "true", "APP_ENV": "development"}, clear=False):
            yanit = istemci.post(
                "/api/v1/veri/dosya-dogrula?dosya_adi=test.csv",
                content="Tarih,Kategori,Gelir,Gider\n2026-01-01,Satış,1000,200\n".encode(),
                headers={"Content-Type": "application/octet-stream"},
            )
        self.assertEqual(yanit.status_code, 200)
        self.assertEqual(yanit.json()["finansal_veri"]["net_kar"], 800)

    def test_musteri_urun_rfm_ve_tahmin_analizleri_uretilir(self):
        satirlar = ["Tarih,Kategori,Gelir,Gider,Müşteri,Ürün"]
        for ay in range(1, 7):
            satirlar.append(f"2026-{ay:02d}-01,Satış,{1000 + ay * 100},0,Müşteri A,Ürün 1")
            satirlar.append(f"2026-{ay:02d}-15,Satış,{500 + ay * 50},0,Müşteri B,Ürün 2")
        sonuc = dosya_dogrula(("\n".join(satirlar) + "\n").encode(), "analiz.csv")

        self.assertEqual(len(sonuc["analizler"]["musteriler"]), 2)
        self.assertEqual(len(sonuc["finansal_veri"]["musteri_cirolari"]), 2)
        self.assertEqual(len(sonuc["analizler"]["urunler"]), 2)
        self.assertEqual(sonuc["analizler"]["tahmin"]["durum"], "hazir")
        self.assertIsNotNone(sonuc["analizler"]["tahmin"]["gecmis_hata_mape"])

    def test_indirilebilir_sablon_yeniden_okunabilir(self):
        sonuc = dosya_dogrula(veri_sablonu_olustur(), "sablon.xlsx")
        self.assertEqual(sonuc["finansal_veri"]["sirket_adi"], "Örnek Şirket A.Ş.")
        self.assertIn("İşlemler", sonuc["dosya"]["sayfalar"])

    def test_mukerrer_fatura_fazla_odeme_hatali_vade_ve_para_birimi_reddedilir(self):
        kitap = Workbook()
        finans = kitap.active
        finans.title = "Finansal_Gorunum"
        finans.append(["Şirket_Adı", "Dönem", "Ciro", "Net_Kâr"])
        finans.append(["Stres Test A.Ş.", "2026", 100_000, 10_000])

        alacak = kitap.create_sheet("Alacaklar")
        alacak.append(["Fatura_ID", "Müşteri_ID", "Müşteri_Adı", "Fatura_Tarihi", "Vade_Tarihi", "Tutar", "Ödenen"])
        alacak.append(["F-1", "M-1", "Müşteri", "2026-01-01", "2026-01-31", 10_000, 2_000])
        alacak.append(["F-1", "M-1", "Müşteri", "2026-02-01", "2026-02-28", 5_000, 0])
        alacak.append(["F-2", "M-2", "Müşteri 2", "2026-03-01", "2026-03-31", 4_000, 5_000])
        alacak.append(["F-3", "M-3", "Müşteri 3", "2026-04-15", "2026-04-01", 4_000, 0])

        borc = kitap.create_sheet("Borc_Servisi")
        borc.append(["Borç_ID", "Alacaklı", "Ödeme_Tarihi", "Anapara", "Faiz", "Para_Birimi"])
        borc.append(["B-1", "Banka", "2026-05-01", 1_000, 100, "XYZ"])

        tampon = io.BytesIO()
        kitap.save(tampon)
        sonuc = dosya_dogrula(tampon.getvalue(), "stres.xlsx")

        self.assertEqual(len(sonuc["gelismis_veri"]["alacak_faturalari"]), 1)
        self.assertEqual(len(sonuc["gelismis_veri"]["borc_servisi"]), 0)
        self.assertTrue(any(hata["kod"] == "mukerrer_kimlik" for hata in sonuc["hatalar"]))
        self.assertGreaterEqual(sonuc["ozet"]["reddedilen_satirlar"], 4)

    def test_sonsuz_sayi_ve_gecersiz_tarih_reddedilir(self):
        icerik = (
            "Tarih,Kategori,Gelir,Gider\n"
            "2026-01-01,Satış,NaN,0\n"
            "31.02.2026,Satış,100,0\n"
            "2026-01-03,Satış,100,0\n"
        ).encode()
        sonuc = dosya_dogrula(icerik, "bozuk.csv")
        self.assertEqual(sonuc["ozet"]["islem_satirlari"], 1)
        self.assertEqual(sonuc["ozet"]["reddedilen_satirlar"], 2)

    def test_dosya_dogrula_kalite_bulgularini_gosterir(self):
        """Response artık cross-field + anomali bulgularını taşımalı."""
        # Bilanço eşitsizliği: 1000 varlık, 400+500=900 pasif
        kitap = Workbook()
        finans = kitap.active
        finans.title = "Finansal_Gorunum"
        finans.append([
            "Şirket_Adı", "Dönem", "Ciro", "Net_Kâr",
            "Toplam_Varlıklar", "Toplam_Yükümlülükler", "Özkaynak",
        ])
        finans.append(["Bilanço Testi", "2026", 100_000, 10_000,
                       1_000_000, 400_000, 500_000])
        tampon = io.BytesIO()
        kitap.save(tampon)

        sonuc = dosya_dogrula(tampon.getvalue(), "bilanco.xlsx")
        vk = sonuc["veri_kalitesi"]
        self.assertIn("tutarlilik_bulgulari", vk)
        self.assertIn("anomali_bulgulari", vk)
        self.assertTrue(any(
            b["kod"] == "bilanco_esitsizligi"
            for b in vk["tutarlilik_bulgulari"]
        ))
        self.assertGreaterEqual(vk["semantik_hata_sayisi"], 1)
        self.assertEqual(vk["semantik_durum"], "hatali")

    def test_dosya_dogrula_sayfa_eslesmesini_ozetler(self):
        """Bilinmeyen sayfalar 'atlanan_sayfalar' listesinde görünmeli."""
        kitap = Workbook()
        finans = kitap.active
        finans.title = "Finansal_Gorunum"
        finans.append(["Şirket_Adı", "Ciro", "Net_Kâr"])
        finans.append(["Test", 100_000, 10_000])
        kitap.create_sheet("Rehber").append(["Nasıl doldurulur"])
        kitap.create_sheet("BenimOzelSayfam").append(["x"])
        tampon = io.BytesIO()
        kitap.save(tampon)

        sonuc = dosya_dogrula(tampon.getvalue(), "sayfa.xlsx")
        dosya = sonuc["dosya"]
        self.assertIn("Finansal_Gorunum", dosya["tanınan_sayfalar"])
        self.assertIn("Rehber", dosya["tanınan_sayfalar"])
        self.assertIn("BenimOzelSayfam", dosya["atlanan_sayfalar"])

    def test_elli_bin_satir_siniri_guvenli_bicimde_uygulanir(self):
        satirlar = ["Tarih,Kategori,Gelir,Gider"]
        satirlar.extend(f"2026-01-01,Satış,{index + 1},0" for index in range(50_005))
        sonuc = dosya_dogrula(("\n".join(satirlar) + "\n").encode(), "hacim.csv")
        self.assertEqual(sonuc["ozet"]["islem_satirlari"], 50_000)
        self.assertTrue(any(hata["kod"] == "satir_siniri" for hata in sonuc["hatalar"]))


if __name__ == "__main__":
    unittest.main()
