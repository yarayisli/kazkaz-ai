import io
import unittest

from openpyxl import load_workbook

from api.models import FinansalGorunum
from api.report_engine import excel_raporu_olustur, pdf_raporu_olustur


class TestYoneticiRaporlari(unittest.TestCase):
    def setUp(self):
        self.veri = FinansalGorunum(
            sirket_adi="Rapor Test A.Ş.",
            sektor="İmalat",
            donem="2026",
            ciro=1_000_000,
            satis_maliyeti=550_000,
            faaliyet_giderleri=220_000,
            net_kar=120_000,
            nakit=180_000,
            kisa_vadeli_borc=140_000,
            uzun_vadeli_borc=200_000,
            alacaklar=160_000,
            borclar=100_000,
            stoklar=90_000,
            ozkaynak=500_000,
            faiz_gideri=25_000,
            vergi_gideri=30_000,
            amortisman=20_000,
            capex=45_000,
            donen_varliklar=430_000,
            toplam_varliklar=1_100_000,
            toplam_yukumlulukler=600_000,
            dagitilmamis_karlar=150_000,
            operasyonel_nakit_akisi=170_000,
        )

    def test_pdf_gecerli_ve_bos_degil(self):
        icerik = pdf_raporu_olustur(self.veri)
        self.assertTrue(icerik.startswith(b"%PDF"))
        self.assertGreater(len(icerik), 2_000)

    def test_excel_beklenen_sayfalari_ve_formulleri_icerir(self):
        icerik = excel_raporu_olustur(self.veri)
        kitap = load_workbook(io.BytesIO(icerik), data_only=False)
        self.assertEqual(kitap.sheetnames, ["Yönetici Özeti", "Metrikler", "Risk ve Aksiyonlar"])
        self.assertEqual(kitap["Yönetici Özeti"]["B2"].value, "Rapor Test A.Ş.")
        formul_kolonlari = [satir[3].value for satir in kitap["Metrikler"].iter_rows(min_row=2)]
        self.assertTrue(any(formul_kolonlari))


if __name__ == "__main__":
    unittest.main()
